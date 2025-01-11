import pytest
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from openpyxl.workbook import Workbook

from ..utilities.excel_handler import ExcelHandler
from ..library.simulator import simulator
from ..pages.navigation_page import NavigationPage

# Constants
BASE_URL = "http://10.179.14.87/logon.htm"
CHROME_DRIVER_PATH = r"C:\Users\SESA791538\Downloads\chromedriver-win64\chromedriver-win64\chromedriver.exe"


@pytest.fixture(scope="session")
def excel_handler():
    """Fixture for Excel workbook handling"""
    wb = Workbook()
    ws = wb.active

    # Initialize headers
    ws['A1'] = 'Function'
    ws['B1'] = 'NMC VALUE'
    ws['C1'] = 'SIMULATOR VALUE'
    ws['D1'] = "OPERATION"
    ws['E1'] = "OPERATION TYPE"

    handler = ExcelHandler(wb)
    yield handler
    handler.save()


@pytest.fixture(scope="class")
def driver():
    """Fixture for WebDriver setup and teardown"""
    service = Service(CHROME_DRIVER_PATH)
    driver = webdriver.Chrome(service=service)
    driver.maximize_window()
    driver.get("http://10.179.14.32/logon.htm")
    time.sleep(10)

    # Login
    driver.find_element(By.XPATH, "//input[@name='login_username']").send_keys("apc")
    driver.find_element(By.XPATH, "//input[@name='login_password']").send_keys("$chne1der@")
    driver.find_element(By.XPATH, "//input[@name='submit']").click()
    time.sleep(10)

    assert "home.htm" in driver.current_url, "Login failed"

    yield driver

    # Logout and quit
    driver.find_element(By.XPATH,"//div[@id='user-area']//a[@title='Log User Off'][normalize-space()='Log Off']").click()
    time.sleep(2)
    driver.quit()


@pytest.fixture(scope="class")
def navigation(driver):
    """Fixture that provides the NavigationPage instance"""
    return NavigationPage(driver)


@pytest.fixture(scope="class")
def simulator_setup():
    """Fixture for simulator setup"""
    sim = simulator.MicrolinkSimulator(simulator.microlink_host_ip,simulator.microlink_sim_path,simulator.microlink_xml_path)
    return sim


# Navigation Fixtures
@pytest.fixture
def navigate_to_home(navigation):
    """Navigate to home page"""
    navigation.navigate_to_home()
    return navigation.driver


@pytest.fixture
def navigate_to_status(navigation):
    """Navigate to status->UPS page"""
    navigation.navigate_to_status()
    return navigation.driver


@pytest.fixture
def navigate_status_outlet_groups(navigation):
    """Navigate to status ->Outlet Groups Page"""
    navigation.navigate_to_status_outlet_groups()
    return navigation.driver


@pytest.fixture
def navigate_status_universalio(navigation):
    """Navigate to status -> Universal IO page"""
    navigation.navigate_to_status_universalio()
    return navigation.driver


@pytest.fixture
def navigate_to_control(navigation):
    """Navigate to control page"""
    navigation.navigate_to_control()
    return navigation.driver


@pytest.fixture
def navigate_to_configuration(navigation):
    """Navigate to configuration page"""
    navigation.navigate_to_configuration()
    return navigation.driver


@pytest.fixture
def navigate_to_about_ups(navigation):
    """Navigate to about page"""
    navigation.navigate_to_about()
    return navigation.driver


@pytest.fixture
def navigate_control_outlet_groups(navigation):
    """Navigate to outlet groups control page"""
    navigation.navigate_to_control_outlet_groups()
    return navigation.driver


@pytest.fixture
def navigate_toFetchMogAndSog(driver):
    """Test navigation and status check for outlet groups"""
    driver.find_element(By.XPATH, "//a[normalize-space()='Status']").click()
    driver.find_element(By.XPATH, "//a[@href='ulstoutg.htm']").click()

    mogStatus = driver.find_element(By.XPATH, "//tbody/tr[2]/td[2]").text
    sogStatus = driver.find_element(By.XPATH, "//tbody/tr[3]/td[2]").text

    assert mogStatus is not None, "MOG status not found"
    assert sogStatus is not None, "SOG status not found"

    return (mogStatus, sogStatus)


@pytest.fixture
def navigate_configuration_shutdown(navigation):
    """Navigate to configuration shutdown page"""
    navigation.navigate_to_configuration_shutdown()
    return navigation.driver


@pytest.fixture
def get_delays(driver, navigate_to_configuration):
    """Get various delay values"""
    return {
        'main_power_off': int(driver.find_element(By.XPATH, "//tbody/tr[2]/td[2]").text),
        'main_power_on': int(driver.find_element(By.XPATH, "//tbody/tr[2]/td[4]").text),
        'sog_power_off': int(driver.find_element(By.CSS_SELECTOR, "tbody tr:nth-child(4) td:nth-child(2)").text),
        'sog_power_on': int(driver.find_element(By.XPATH, "//tbody/tr[4]/td[4]").text),
        'reboot_duration': int(driver.find_element(By.XPATH, "//tbody/tr[2]/td[3]").text)
    }