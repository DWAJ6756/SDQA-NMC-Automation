import os
import telnetlib
import time
import lovely_logger as log
import subprocess
from . import XML
from . import Microlink_functions
import openpyxl
import os



microlink_host_ip = "localhost"
# microlink_sim_path = "C:\\Users\\Administrator\\Desktop\\Microlink simulator\\ulSim 4.0.0.7\\ulSim 4.0.0.7\\ulSim 4.0.0.7"
# microlink_xml_path = "C:\\Users\\Administrator\\Desktop\\automation\\ID_1027\\ID_1027\\Latest Firmware\\SMT1027-45_xml (1)\\ulsim_cfg_com_SMT_230V_COM1_9600_factory_mode.xml"
# excel_file = 'C:/Users/Administrator/Desktop/automation/simulator/output.xlsx'
microlink_sim_path=None
microlink_xml_path=None
excel_file=None



current_directory = os.getcwd()
output_filename = "microlink_output.txt"
file_path = os.path.join(current_directory, output_filename)



days = int()
Hex_values = ""
Binary_values = ""
Ascii_values = ""
Decimal_values = ""
String_values = ""
Manufacture_date = ""
String_values = ""
class MicrolinkSimulator:
    global file_path
    def __init__(self, microlink_host_ip, microlink_sim_path, microlink_xml_path):
        self.microlink_host_ip = microlink_host_ip
        self.microlink_sim_path = microlink_sim_path
        self.microlink_xml_path = microlink_xml_path
        self.current_dir = os.getcwd()
        log.info(f"Current directory: {self.current_dir}")
        self.tn = None
        self.connect_to_simulator()

    global excel_file
    def connect_to_simulator(self):
        try:
            self.tn = telnetlib.Telnet(self.microlink_host_ip, 8080, 10)
            log.info("Connected to Microlink on port 8080")
        except Exception as ex:
            log.error(f"Failed to connect on port 8080: {ex}")
            self.launch_simulator()
            self.retry_connection()

    def launch_simulator(self):
        try:
            os.chdir(self.microlink_sim_path)
            command = f"cmd /K ulSim-TreeView.bat {self.microlink_xml_path}"
            subprocess.Popen(command, shell=True)
            time.sleep(20)
            log.info('Microlink launched')
        except Exception as ex:
            log.error(f"Failed to launch simulator: {ex}")
        finally:
            os.chdir(self.current_dir)



    def retry_connection(self):
        try:
            self.tn = telnetlib.Telnet(self.microlink_host_ip, 8080, 10)
            log.info("Reconnected to Microlink on port 8080")
        except Exception as ex:
            log.error(f"Retry failed on port 8080: {ex}")
            self.try_alternate_port()

    def try_alternate_port(self):
        try:
            self.tn = telnetlib.Telnet(self.microlink_host_ip, 8081, 10)
            log.info("Connected to Microlink on port 8081")
        except Exception as ex:
            log.error(f"Failed to connect on port 8081: {ex}")
            self.launch_simulator()
            self.retry_connection()

    def UPSStatus_BF(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        def write_to_excel(com_to_send, UPS_Status):
            if not os.path.exists(excel_file):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.append(['Command', 'Response'])
            else:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

            sheet.append([com_to_send, UPS_Status])
            workbook.save(excel_file)


        com_to_send = "getusage Power1:UPSSystem.UPSStatus_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    # print(f"UPS_Status: {XML.Binary_values}")
                    combined_status = []
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global UPS_status
                            UPS_status = Microlink_functions.UPS_Status[i]
                            combined_status.append(UPS_status)
                            # print(f"UPS Status+++++: {UPS_status}, ", end="")

                    for i in range(len(combined_status)):
                        print(combined_status[i]+"\n")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)

        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    # def UPSStatus_BF(self):
    #
    #     com_to_send = "getusage Power1:UPSSystem.UPSStatus_BF"
    #     print(com_to_send)
    #     log.info(f"Command to send: {com_to_send}")
    #
    #
    #     if self.tn is None:
    #         self.connect_to_simulator()
    #
    #     self.tn.write(com_to_send.encode('ascii') + b"\n")
    #     log.info("Command sent successfully")
    #
    #     time.sleep(2)
    #     response = self.tn.read_very_eager().decode('ascii')
    #     # log.info(f"Raw response: {response}")
    #
    #     print(response)


    def SerialNumber_STR(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"
        excel_file = 'output.xlsx'

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        def write_to_excel(com_to_send, UPS_SerialNumber):
            if not os.path.exists(excel_file):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.append(['Command', 'Response'])
            else:
                try:
                    workbook = openpyxl.load_workbook(excel_file)
                    sheet = workbook.active
                except Exception as e:
                    log.error(f"Error opening Excel file: {e}")
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    sheet.append(['Command', 'Response'])

            sheet.append([com_to_send, UPS_SerialNumber])
            workbook.save(excel_file)

        com_to_send = "getusage Power1:UPSSystem.ManufacturerData.SerialNumber_STR"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    combined_serial_numbers = []
                    serial_number = XML.String_values
                    combined_serial_numbers.append(serial_number)
                    print(f"UPS SerialNumber: {serial_number}")
                    combined_serial_numbers_str = "".join(combined_serial_numbers)
                    write_to_excel(com_to_send, combined_serial_numbers_str)

                    return serial_number
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)

        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def ManufactureDate(self):
        global days


        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.ManufacturerData.Date"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    days = XML.days
                    XML.calculate_manufacture_date(days)
                    global UPS_ManufactureDate
                    UPS_ManufactureDate = XML.Manufacture_date
                    print(f"UPS_Manufacture_date: {XML.Manufacture_date}")
                    return UPS_ManufactureDate
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSModel_STR(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.ManufacturerData.Model_STR"
        # print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")
                print(response)

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global UPS_Model
                    UPS_Model = XML.String_values
                    print(f"UPS_Model: {XML.String_values}")
                    UPS_Model = UPS_Model.replace("\x01", " ")
                    return UPS_Model

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSKU_STR(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.ManufacturerData.SKU_STR"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global UPS_SKU
                    UPS_SKU = XML.String_values
                    print(f"UPS_SKU: {XML.String_values}")
                    return UPS_SKU
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def Battery_Insatllation_date(self):
        global days


        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.BatterySystem.DateSetting"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    days = XML.days
                    XML.calculate_manufacture_date(days)
                    global Battery_Installation_Date
                    Battery_Installation_Date = XML.Manufacture_date
                    print(f"Battery_Installation_date: {XML.Manufacture_date}")
                    return Battery_Installation_Date
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def Battery_replace_AlarmNotificationSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.BatterySystem.Power2:AlarmNotificationSetting_D_EN"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global Battery_Replace_AlarmNotificationSetting
                    Battery_Replace_AlarmNotificationSetting = XML.Decimal_values
                    print(f"Battery_replace_AlarmNotificationSetting: {XML.Decimal_values}")
                    return Battery_Replace_AlarmNotificationSetting

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def Battery_replace_AlarmReminderInterval(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.BatterySystem.Power2:AlarmReminderIntervalSetting_D_EN"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global Battery_replace_AlarmReminderInterval
                    Battery_replace_AlarmReminderInterval = XML.Decimal_values
                    print(f"Battery_replace_AlarmReminderInterval: {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")


    def UPSBatterySKU(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.BatterySystem.SKU_STR"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global Battery_SKU
                    Battery_SKU = XML.String_values
                    print(f"Battery_part_no: {XML.String_values}")
                    return Battery_SKU
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def BatteryTest_IntervalSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.BatterySystem.BatteryTestIntervalSetting_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"BatteryTest_IntervalSetting: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global BatteryTest_IntervalSetting
                            BatteryTest_IntervalSetting = Microlink_functions.Battery_test_IntervalSetting[i]
                            print(f"BatteryTest_IntervalSetting: {Microlink_functions.Battery_test_IntervalSetting[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def Battery_LowRuntimeWarningSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.BatterySystem.LowRuntimeWarningSetting"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global Battery_LowRuntimeWarning
                    Battery_LowRuntimeWarning = XML.Decimal_values
                    print(f"Battery_Low Runtime Remaining: {XML.Decimal_values}")
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def Battery_LifeTimeStatus(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.BatterySystem.LifeTimeStatus_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"Battery Lifetime Status: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global Battery_LifetimeStatus
                            Battery_LifetimeStatus = Microlink_functions.Battery_LifeTime_Status[i]
                            print(f"Battery Life Time Status: {Microlink_functions.Battery_LifeTime_Status[i]}")
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)
        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def Battery_Volatge(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.BatterySystem.VoltageDC"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"Battery_Voltage: {round(XML.Decimal_values / 32, 1)}")
                    global Battery_Voltage
                    Battery_Voltage = round(XML.Decimal_values / 32, 1)
                    return Battery_Voltage
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)
        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def Battery_StateOfCharge(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.BatterySystem.StateOfCharge_Pct"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()
                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")
                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")
                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"Battery_State Of Charge: {round(XML.Decimal_values / 512, 1)}")
                    global Battery_StateOfCharge
                    Battery_StateOfCharge = round(XML.Decimal_values / 512, 1)
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def Battery_ReplaceTest_Status(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.BatterySystem.ReplaceBatteryTestStatus_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()
                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")
                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")
                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"Battery Self Test Status: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global Battery_ReplaceTest_Status
                            Battery_ReplaceTest_Status = Microlink_functions.Battery_ReplaceTest_Status[i]
                            print(f"Battery Self Test Status: {Microlink_functions.Battery_ReplaceTest_Status[i]}")
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)
        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def Battery_RuntimeCalibrationTest_Status(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.BatterySystem.RunTimeCalibrationStatus_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()
                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")
                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")
                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"Battery Runtime Calibration Test Status: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global Battery_RuntimeCalibration_Status
                            Battery_RuntimeCalibration_Status = Microlink_functions.Battery_RuntimeCalibrationTest_Status[i]
                            print(f"Battery Runtime Calibration Test Status: {Microlink_functions.Battery_RuntimeCalibrationTest_Status[i]}")
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)
        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def Battery_RuntimeRemaining(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.BatterySystem.RunTimeRemaining"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()
                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")
                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")
                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"Battery Runtime Remaining: {round(XML.Decimal_values / 60, 1)}")
                    global Battery_RuntimeRemaining
                    Battery_RuntimeRemaining = round(XML.Decimal_values / 60, 1)
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def Battery_Temperature(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")
        com_to_send = "getusage Power1:UPSSystem.BatterySystem.Temperature"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0
        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")
                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")
                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"Battery_Temperature: {round(XML.Decimal_values / 128, 1)}")
                    global Battery_Temperature
                    Battery_Temperature = round(XML.Decimal_values / 128, 1)
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)
        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def OutputSystem_UpperAcceptableVoltageSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.OutputSystem.UpperAcceptableVoltageSetting"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0
        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()
                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")
                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")
                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"Output Voltage Upper Acceptable limit: {XML.Decimal_values}")
                    global OutputSystem_UpperAcceptableVoltageSetting
                    OutputSystem_UpperAcceptableVoltageSetting = XML.Decimal_values
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def OutputSystem_LowerAcceptableVoltageSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.OutputSystem.LowerAcceptableVoltageSetting"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"Output Voltage Lower Acceptable limit: {XML.Decimal_values}")
                    global OutputSystem_LowerAcceptableVoltageSetting
                    OutputSystem_LowerAcceptableVoltageSetting = XML.Decimal_values
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)
        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def OutputSystem_UPS_Sensitivity(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.OutputSystem.Power2:SensitivitySetting_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Sensitivity: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global UPS_Sensitivity
                            UPS_Sensitivity = Microlink_functions.UPS_Sensitivity[i]
                            print(f"UPS Sensitivity: {Microlink_functions.UPS_Sensitivity[i]}")
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)
        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def OutputSystem_Energy(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")
        com_to_send = "getusage Power1:UPSSystem.OutputSystem.Power2:Energy_4B"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0
        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()
                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")
                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")
                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Output Energy: {XML.Decimal_values / 1000}")
                    global OutputSystem_Energy
                    OutputSystem_Energy = XML.Decimal_values / 1000
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def Output_Rated_ApparentPower(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.OutputSystem.ApparentPowerRating"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Output Rated Apparent Power: {XML.Decimal_values}")
                    global Output_Rated_ApparentPower
                    Output_Rated_ApparentPower = XML.Decimal_values
                    return Output_Rated_ApparentPower
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)
        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def Output_Rated_RealPower(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.OutputSystem.RealPowerRating"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Output Rated Real Power: {XML.Decimal_values}")
                    global Output_Rated_RealPower
                    Output_Rated_RealPower = XML.Decimal_values
                    return Output_Rated_RealPower
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def OutputSystem_VoltageACSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")
        com_to_send = "getusage Power1:UPSSystem.OutputSystem.VoltageACSetting_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0
        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()
                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")
                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Output Voltage Setting: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global UPS_Output_Voltage_Setting
                            UPS_Output_Voltage_Setting = Microlink_functions.Output_VoltageSetting[i]
                            print(f"UPS Output Voltage Setting: {Microlink_functions.Output_VoltageSetting[i]}")
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)
        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def OutputSystem_VoltageAC(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")
        com_to_send = "getusage Power1:UPSSystem.OutputSystem.VoltageAC"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0
        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()
                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")
                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")
                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Output Voltage: {round(XML.Decimal_values / 64, 1)}")
                    global OutputSystem_VoltageAC
                    OutputSystem_VoltageAC = round(XML.Decimal_values / 64, 1)
                    return OutputSystem_VoltageAC
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def OutputSystem_CurrentAC(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"
        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")
        com_to_send = "getusage Power1:UPSSystem.OutputSystem.CurrentAC"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0
        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Output Current: {XML.Decimal_values}")
                    global OutputSystem_CurrentAC
                    OutputSystem_CurrentAC = XML.Decimal_values
                    return OutputSystem_CurrentAC
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def OutputSystem_Frequency(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"
        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")
        com_to_send = "getusage Power1:UPSSystem.OutputSystem.Frequency"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0
        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()
                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")
                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")
                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Output Frequency:: {round(XML.Decimal_values / 128, 1)}")
                    global OutputSystem_Frequency
                    OutputSystem_Frequency = round(XML.Decimal_values / 128, 1)
                    return OutputSystem_Frequency
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def OutputSystem_Actual_ApparentPower(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.OutputSystem.ApparentPower_Pct"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Output Actual Apparent Power: {XML.Decimal_values}")
                    global OutputSystem_Actual_ApparentPower
                    OutputSystem_Actual_ApparentPower = XML.Decimal_values
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)
        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def OutputSystem_Actual_RealPower(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.OutputSystem.RealPower_Pct"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Actual Real Power: {XML.Decimal_values}")
                    global OutputSystem_Actual_RealPower
                    OutputSystem_Actual_RealPower = XML.Decimal_values
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_UnswitchedOutletGroup_Name(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.UnswitchedOutletGroup.Name_STR"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global UPS_MOGName
                    UPS_MOGName = XML.String_values
                    print(f"UPS Unswitched Outlet Group Name : {XML.String_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")


    def UPSSystem_UnswitchedOutletGroup_TurnOnCountdownSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.UnswitchedOutletGroup.TurnOnCountdownSetting_EN"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Unswitched Outlet Group Turn On countdown: {XML.Decimal_values}")
                    global UPS_MOG_TurnONCountdownSetting
                    UPS_MOG_TurnONCountdownSetting = XML.Decimal_values

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_UnswitchedOutletGroup_TurnOffCountdownSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.UnswitchedOutletGroup.TurnOffCountdownSetting_EN"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Unswitched Outlet Group Turn Off countdown: {XML.Decimal_values}")
                    global UPS_MOG_TurnOFFCountdown
                    UPS_MOG_TurnOFFCountdown = XML.Decimal_values

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)

        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_UnswitchedOutletGroup_StayOffCountdownSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")
        com_to_send = "getusage Power1:UPSSystem.UnswitchedOutletGroup.StayOffCountdownSetting_4B"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Unswitched Outlet Group Stay Off countdown: {XML.Decimal_values}")
                    global UPS_MOG_StayOFFCountdown
                    UPS_MOG_StayOFFCountdown = XML.Decimal_values
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_UnSwitchedOutletGroup_LoadShedConfigSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.UnswitchedOutletGroup.Power2:LoadShedConfigSetting_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS MOG Load Shed Config Setting: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global UPS_MOG_LoadShedConfig
                            UPS_MOG_LoadShedConfig = Microlink_functions.UnSwitched_OutletGroup_LoadShedConfigSetting[i]
                            print(f"UPS Unswitched Outlet Group LoadShed Time On Battery Setting: {Microlink_functions.UnSwitched_OutletGroup_LoadShedConfigSetting[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")


    def UPSSystem_UnswitchedOutletGroup_LoadShedRuntimeRemaining(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.UnswitchedOutletGroup.Power2:LoadShedRunTimeRemainingSetting"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Unswitched Outlet Group LoadShed Runtime Remaining Setting : {XML.Decimal_values}")
                    global UPS_MOG_LoadShedRuntimeRemaining
                    UPS_MOG_LoadShedRuntimeRemaining = XML.Decimal_values

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_UnswitchedOutletGroup_LoadShedTimeOnBattery(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.UnswitchedOutletGroup.Power2:LoadShedTimeOnBatterySetting"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Unswitched Outlet Group LoadShed Time On Battery Setting : {XML.Decimal_values}")
                    global UPS_MOG_LoadShedTimeOnBattery
                    UPS_MOG_LoadShedTimeOnBattery = XML.Decimal_values

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_UnswitchedOutletGroup_MinimumReturnRuntime(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.UnswitchedOutletGroup.MinimumReturnRuntimeSetting"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Unswitched Outlet Group Minimum Return Runtime Setting : {XML.Decimal_values}")
                    global UPS_MOG_MinimumReturnRuntime
                    UPS_MOG_MinimumReturnRuntime = XML.Decimal_values

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")


    def UPSSystem_UnSwitchedOutletGroup_Status(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.UnswitchedOutletGroup.OutletStatus_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS MOG Status: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global UPS_MOG_Status
                            UPS_MOG_Status = Microlink_functions.UnSwitched_OutletGroup_Status[i]
                            print(f"UPS Unswitched Outlet Group Status: "+UPS_MOG_Status)
                    return UPS_MOG_Status

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")



    def UPSSystem_UserInterfaceSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.UserInterfaceSystem.UserInterfaceSetting_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS User Interface Setting: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global UPS_UserInterfaceSetting
                            UPS_UserInterfaceSetting = Microlink_functions.UPS_UserInterfaceSetting[i]
                            print(f"UPS User Interface setting: {Microlink_functions.UPS_UserInterfaceSetting[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_UserInterfaceCommand(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.UserInterfaceSystem.UserInterfaceCommand_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS User Interface Command: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global UPS_UserInterfaceCommand
                            UPS_UserInterfaceCommand = Microlink_functions.UPS_UserInterfaceCommand[i]
                            print(f"UPS User Interface Command: {Microlink_functions.UPS_UserInterfaceCommand[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_UserInterfaceStatus(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.UserInterfaceSystem.UserInterfaceStatus_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS User Interface Status: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global UPS_UserInterfaceStatus
                            UPS_UserInterfaceStatus = Microlink_functions.UPS_UserInterfaceStatus[i]
                            print(f"UPS User Interface Status: {Microlink_functions.UPS_UserInterfaceStatus[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_FileTransferSourceSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.CommunicationSystem.Power2:FileTransferSourcesSetting_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS File Transfer Source Setting: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global UPS_FileTransferSourceSetting
                            UPS_FileTransferSourceSetting = Microlink_functions.UPS_FileTransferSourceSetting[i]
                            print(f"UPS File Transfer Source Setting: {Microlink_functions.UPS_FileTransferSourceSetting[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SmartSlotSKU(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.Management1:SmartSlot.ManufacturerData.SKU_STR"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global UPS_SmartSlotSKU
                    UPS_SmartSlotSKU = XML.Ascii_values
                    print(f"UPS Smart Slot SKU : {XML.Ascii_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)
        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SmartSlotHWversion(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.Management1:SmartSlot.ManufacturerData.HWVersion_STR"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Smart Slot Hardware Version : {XML.Ascii_values}")
                    global UPS_SmartSlotHardwareVersion
                    UPS_SmartSlotHardwareVersion = XML.Ascii_values

                    return UPS_SmartSlotHardwareVersion

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SmartSlotProbe0Name(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.Management1:SmartSlot.Probe[00].Name_STR"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Smart Slot Probe 0 Name : {XML.String_values}")
                    global UPS_SmartSLot_Probe0_Name
                    UPS_SmartSLot_Probe0_Name = XML.String_values

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_Smartslot_Probe0Type(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.Management1:SmartSlot.Probe[00].ProbeType_EN"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Smartslot Probe 0 Type: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global UPS_Smartslot_Probe0_Type
                            UPS_Smartslot_Probe0_Type = Microlink_functions.UPS_Smartslot_Probe0Type[i]
                            print(f"UPS Smartslot Probe 0 Type: {Microlink_functions.UPS_Smartslot_Probe0Type[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SmartSlotProbe0Temperature(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.Management1:SmartSlot.Probe[00].Temperature"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()
                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")
                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")
                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)

                    hexa_value=XML.Decimal_values
                    temp=XML.convert_decimal_to_custom_value_temp(hexa_value)
                    print(temp)

                    global UPS_SmartSlot_Probe0_Temperature
                    UPS_SmartSlot_Probe0_Temperature = XML.Decimal_values
                    return temp


                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SmartSlotProbe0Humidity(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.Management1:SmartSlot.Probe[00].RelativeHumidity_Pct"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")
                # print("Response"+response)

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    decimal_Value=XML.Decimal_values
                    simulator_decimal_value=XML.convert_decimal_to_custom_value(decimal_Value)
#                     print(f"UPS Smart Slot Probe 0 Humidity : {simulator_decimal_value}")
# # ------------------------------------------------------------------------------------------------
#                     global UPS_SmartSlot_Probe0_Humidity
#                     UPS_SmartSlot_Probe0_Humidity = XML.Decimal_values
                    return simulator_decimal_value
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        # if retries == max_retries:
        #     log.error("Max retries reached. Command failed.")

    def UPSSystem_Smartslot_Probe0Status(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.Management1:SmartSlot.Probe[00].ProbeStatus_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Smartslot Probe 0 Status: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global UPS_Smartslot_Probe0_Status
                            UPS_Smartslot_Probe0_Status = Microlink_functions.UPS_Smartslot_Probe0Status[i]
                            print(f"UPS Smartslot Probe 0 Status: {Microlink_functions.UPS_Smartslot_Probe0Status[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_Smartslot_Probe0UniversalStatus(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.Management1:SmartSlot.Probe[00].UniversalStatus_EN"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Smartslot Probe 0 Universal Status: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global UPS_Smartslot_Probe0_UniversalStatus
                            UPS_Smartslot_Probe0_UniversalStatus = Microlink_functions.UPS_Smartslot_Probe0UniversalStatus[i]
                            print(f"UPS Smartslot Probe 0 Universal Status: {Microlink_functions.UPS_Smartslot_Probe0UniversalStatus[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SmartSlotProbe1Name(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.Management1:SmartSlot.Probe[01].Name_STR"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global UPS_SmartSlot_Probe1Name
                    UPS_SmartSlot_Probe1Name = XML.String_values
                    print(f"UPS Smart Slot Probe 1 Name : {XML.String_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_Smartslot_Probe1Type(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.Management1:SmartSlot.Probe[01].ProbeType_EN"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Smartslot Probe 1 Type: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global UPS_SmartSlot_Probe1Type
                            UPS_SmartSlot_Probe1Type = Microlink_functions.UPS_Smartslot_Probe1Type[i]
                            print(f"UPS Smartslot Probe 1 Type: {Microlink_functions.UPS_Smartslot_Probe1Type[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SmartSlotProbe1Temperature(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.Management1:SmartSlot.Probe[01].Temperature"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global UPS_SmartSlot_Probe1_Temperature
                    UPS_SmartSlot_Probe1_Temperature = XML.Decimal_values
                    print(f"UPS Smart Slot Probe 1 Temperature : {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SmartSlotProbe1Humidity(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.Management1:SmartSlot.Probe[01].RelativeHumidity_Pct"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global UPS_SmartSlot_Probe1_Humidity
                    UPS_SmartSlot_Probe1_Humidity = XML.Decimal_values
                    print(f"UPS Smart Slot Probe 1 Humidity : {XML.Decimal_values}")
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_Smartslot_Probe1Status(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.Management1:SmartSlot.Probe[01].ProbeStatus_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Smartslot Probe 1 Status: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global UPS_Smartslot_Probe1Status
                            UPS_Smartslot_Probe1Status = Microlink_functions.UPS_Smartslot_Probe1Status[i]
                            print(f"UPS Smartslot Probe 1 Status: {Microlink_functions.UPS_Smartslot_Probe1Status[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_Smartslot_Probe1UniversalStatus(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.Management1:SmartSlot.Probe[01].UniversalStatus_EN"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Smartslot Probe 1 Universal Status: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global UPS_Smartslot_Probe1_UniversalStatus
                            UPS_Smartslot_Probe1_UniversalStatus = Microlink_functions.UPS_Smartslot_Probe1UniversalStatus[i]
                            print(f"UPS Smartslot Probe 1 Universal Status: {Microlink_functions.UPS_Smartslot_Probe1UniversalStatus[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_Name(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.Name_STR"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global UPS_Name
                    UPS_Name = XML.String_values
                    print(f"UPS Name : {XML.String_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_HighEfficiencyMode(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.Power2:AllowedOperatingModeSetting_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS High Efficiency Mode: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global High_Efficiency_Mode
                            High_Efficiency_Mode = Microlink_functions.UPS_GreenMode[i]
                            print(f"UPS High Efficiency mode: {Microlink_functions.UPS_GreenMode[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_PowerQualitySetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.Power2:PowerQualitySetting_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Power Quality Setting: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global Power_Quality_Setting
                            Power_Quality_Setting = Microlink_functions.UPS_PowerQualitySetting[i]
                            print(f"UPS Power Quality Setting: {Microlink_functions.UPS_PowerQualitySetting[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_PowerSystemError(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.PowerSystemError_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Power System Error: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global Power_System_Error
                            Power_System_Error = Microlink_functions.UPS_PowerSystemError[i]
                            print(f"UPS Power System Error: {Microlink_functions.UPS_PowerSystemError[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_GeneralError(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.GeneralError_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS General Error: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global General_Error
                            General_Error = Microlink_functions.UPS_GeneralError[i]
                            print(f"UPS General Error: {Microlink_functions.UPS_GeneralError[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_BatterySystemError(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.BatterySystemError_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Battery System Error: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global Battery_systemError
                            Battery_SystemError = Microlink_functions.UPS_BatterySystemError[i]
                            print(f"UPS Battery system Error: {Microlink_functions.UPS_BatterySystemError[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_Command(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.UPSCommand_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Command: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global UPS_Command
                            UPS_Command = Microlink_functions.UPS_Command[i]
                            print(f"UPS Command: {Microlink_functions.UPS_Command[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_OutletCommand(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.OutletCommand_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Outlet Command: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global UPS_OutletCommand
                            UPS_OutletCommand = Microlink_functions.UPS_OutletCommand[i]
                            print(f"UPS Outlet Command: {Microlink_functions.UPS_OutletCommand[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_StatusChangeCause(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.UPSStatusChangeCause_EN"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Status Change Cause: {XML.Decimal_values}")
                    if 0 <= XML.Decimal_values < len(Microlink_functions.UPS_StatusChangeCause):
                        global UPSSystem_StatusChangeCause
                        UPSSystem_StatusChangeCause = Microlink_functions.UPS_StatusChangeCause[XML.Decimal_values]
                        print(f"UPS Status Change Cause: {Microlink_functions.UPS_StatusChangeCause[XML.Decimal_values]}")
                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")


    def UPSSystem_SwitchedOutletGroup1_Name(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[00].Name_STR"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG1_Name
                    SOG1_Name = XML.String_values
                    print(f"UPS switched Outlet Group1 Name : {XML.String_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup2_Name(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[01].Name_STR"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG2_Name
                    SOG2_Name = XML.String_values
                    print(f"UPS switched Outlet Group2 Name : {XML.String_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup3_Name(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[02].Name_STR"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG3_Name
                    SOG3_Name = XML.String_values
                    print(f"UPS switched Outlet Group3 Name : {XML.String_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup1_LowRuntimeWarningSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[00].LowRuntimeWarningSetting"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG1_LowRuntimeWarning
                    SOG1_LowRuntimeWarning = XML.Decimal_values
                    print(f"UPS Switched Outlet Group 1 Low Runtime Warning Setting: {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup2_LowRuntimeWarningSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[01].LowRuntimeWarningSetting"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG2_LowRuntimeWarning
                    SOG2_LowRuntimeWarning = XML.Decimal_values
                    print(f"UPS Switched Outlet Group 2 Low Runtime Warning Setting: {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup3_LowRuntimeWarningSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[02].LowRuntimeWarningSetting"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG3_LowRuntimeWarning
                    SOG3_LowRuntimeWarning = XML.Decimal_values
                    print(f"UPS Switched Outlet Group 3 Low Runtime Warning Setting: {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")


    def UPSSystem_SwitchedOutletGroup1_TurnOnCountdownSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[00].TurnOnCountdownSetting_EN"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG1_TurnON_Countdown
                    SOG1_TurnON_Countdown = XML.Decimal_values
                    print(f"UPS Switched Outlet Group1 Turn On countdown: {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup2_TurnOnCountdownSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[01].TurnOnCountdownSetting_EN"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG2_TurnON_Countdown
                    SOG2_TurnON_Countdown = XML.Decimal_values
                    print(f"UPS Switched Outlet Group2 Turn On countdown: {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)

        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup3_TurnOnCountdownSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[02].TurnOnCountdownSetting_EN"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG3_TurnON_Countdown
                    SOG3_TurnON_Countdown = XML.Decimal_values
                    print(f"UPS Switched Outlet Group3 Turn On countdown: {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)

        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup1_TurnOffCountdownSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[00].TurnOffCountdownSetting_EN"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG1_TurnOFFCountdown
                    SOG1_TurnOFFCountdown = XML.Decimal_values
                    print(f"UPS Switched Outlet Group 1 Turn Off countdown: {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup2_TurnOffCountdownSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[01].TurnOffCountdownSetting_EN"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG2_TurnOFFCountdown
                    SOG2_TurnOFFCountdown = XML.Decimal_values
                    print(f"UPS Switched Outlet Group 2 Turn Off countdown: {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)

        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup3_TurnOffCountdownSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[02].TurnOffCountdownSetting_EN"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG3_TurnOFFCountdown
                    SOG3_TurnOFFCountdown = XML.Decimal_values
                    print(f"UPS Switched Outlet Group 3 Turn Off countdown: {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)

        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup1_StayOffCountdownSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[00].StayOffCountdownSetting_4B"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG1_StayOFFCountdown
                    SOG1_StayOFFCountdown = XML.Decimal_values
                    print(f"UPS Switched Outlet Group 1 Stay Off countdown: {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup2_StayOffCountdownSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[01].StayOffCountdownSetting_4B"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG2_StayOFFCountdown
                    SOG2_StayOFFCountdown = XML.Decimal_values
                    print(f"UPS Switched Outlet Group 2 Stay Off countdown: {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup3_StayOffCountdownSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[02].StayOffCountdownSetting_4B"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG3_StayOFFCountdown
                    SOG3_StayOFFCountdown = XML.Decimal_values
                    print(f"UPS Switched Outlet Group 3 Stay Off countdown: {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)

        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup1_LoadShedConfigSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[00].Power2:LoadShedConfigSetting_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS MOG Load Shed Config Setting: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global SOG1_LoadShedConfig
                            SOG1_LoadShedConfig = Microlink_functions.UnSwitched_OutletGroup_LoadShedConfigSetting[i]
                            print(f"UPS Switched Outlet Group 1 LoadShed Config Setting: {Microlink_functions.UnSwitched_OutletGroup_LoadShedConfigSetting[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup2_LoadShedConfigSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[01].Power2:LoadShedConfigSetting_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS MOG Load Shed Config Setting: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global SOG2_LoadShedConfig
                            SOG2_LoadShedConfig = Microlink_functions.UnSwitched_OutletGroup_LoadShedConfigSetting[i]
                            print(f"UPS Switched Outlet Group 2 LoadShed Config Setting: {Microlink_functions.UnSwitched_OutletGroup_LoadShedConfigSetting[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup3_LoadShedConfigSetting(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[02].Power2:LoadShedConfigSetting_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS MOG Load Shed Config Setting: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global SOG3_LoadShedConfig
                            SOG3_LoadShedConfig = Microlink_functions.UnSwitched_OutletGroup_LoadShedConfigSetting[i]
                            print(f"UPS Switched Outlet Group 3 LoadShed Config Setting: {Microlink_functions.UnSwitched_OutletGroup_LoadShedConfigSetting[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup1_LoadShedRuntimeRemaining(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[00].Power2:LoadShedRunTimeRemainingSetting"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG1_LoadShedRuntimeRemaining
                    SOG1_LoadShedRuntimeRemaining = XML.Decimal_values
                    print(f"UPS Switched Outlet Group 1 LoadShed Runtime Remaining Setting : {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup2_LoadShedRuntimeRemaining(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[01].Power2:LoadShedRunTimeRemainingSetting"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG2_LoadShedRuntimeRemaining
                    SOG2_LoadShedRuntimeRemaining = XML.Decimal_values
                    print(f"UPS Switched Outlet Group 2 LoadShed Runtime Remaining Setting : {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup3_LoadShedRuntimeRemaining(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[02].Power2:LoadShedRunTimeRemainingSetting"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG3_LoadShedRuntimeRemaining
                    SOG3_LoadShedRuntimeRemaining = XML.Decimal_values
                    print(f"UPS Switched Outlet Group 3 LoadShed Runtime Remaining Setting : {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup1_LoadShedTimeOnBattery(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[00].Power2:LoadShedTimeOnBatterySetting"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG1_LoadShedTimeONBattery
                    SOG1_LoadShedTimeONBattery = XML.Decimal_values
                    print(f"UPS Switched Outlet Group 1 LoadShed Time On Battery Setting : {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup2_LoadShedTimeOnBattery(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[01].Power2:LoadShedTimeOnBatterySetting"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG2_LoadShedTimeONBattery
                    SOG2_LoadShedTimeONBattery = XML.Decimal_values
                    print(f"UPS Switched Outlet Group 2 LoadShed Time On Battery Setting : {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup3_LoadShedTimeOnBattery(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[02].Power2:LoadShedTimeOnBatterySetting"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG3_LoadShedTimeONBattery
                    SOG3_LoadShedTimeONBattery = XML.Decimal_values
                    print(f"UPS Switched Outlet Group 3 LoadShed Time On Battery Setting : {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup1_MinimumReturnRuntime(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[00].MinimumReturnRuntimeSetting"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG1_MinimumReturnRuntime
                    SOG1_MinimumReturnRuntime = XML.Decimal_values
                    print(f"UPS Switched Outlet Group 1 Minimum Return Runtime Setting : {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup2_MinimumReturnRuntime(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[01].MinimumReturnRuntimeSetting"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG2_MinimumReturnRuntime
                    SOG2_MinimumReturnRuntime = XML.Decimal_values
                    print(f"UPS Switched Outlet Group 2 Minimum Return Runtime Setting : {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup3_MinimumReturnRuntime(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[02].MinimumReturnRuntimeSetting"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    global SOG3_MinimumReturnRuntime
                    SOG3_MinimumReturnRuntime = XML.Decimal_values
                    print(f"UPS Switched Outlet Group 3 Minimum Return Runtime Setting : {XML.Decimal_values}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup1_Status(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[00].OutletStatus_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    # print(f"UPS Outlet Switched Group 1 Status: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global SOG1_Status
                            SOG1_Status = Microlink_functions.UnSwitched_OutletGroup_Status[i]
                            print(f"UPS Switched Outlet Group 1 Status:"+SOG1_Status)
                    return SOG1_Status

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup2_Status(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[01].OutletStatus_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    # print(f"UPS Outlet Switched Group 2 Status: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global SOG2_Status
                            SOG2_Status = Microlink_functions.UnSwitched_OutletGroup_Status[i]
                            print(f"UPS Switched Outlet Group 2 Status:"+SOG2_Status)

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")

    def UPSSystem_SwitchedOutletGroup3_Status(self):

        XML.empty_file(file_path)
        output_file = "microlink_output.txt"

        def write_to_file(data):
            with open(output_file, 'a') as file:
                file.write(data + "\n")

        com_to_send = "getusage Power1:UPSSystem.SwitchedOutletGroup[02].OutletStatus_BF"
        print(com_to_send)
        log.info(f"Command to send: {com_to_send}")
        max_retries = 3
        retries = 0

        while retries < max_retries:
            try:
                if self.tn is None:
                    self.connect_to_simulator()

                self.tn.write(com_to_send.encode('ascii') + b"\n")
                log.info("Command sent successfully")

                time.sleep(2)
                response = self.tn.read_very_eager().decode('ascii')
                log.info(f"Raw response: {response}")

                if "ERROR" in response:
                    log.error(f"Received error response: {response}")
                    retries += 1
                    log.info(f"Retrying command... Attempt {retries}")
                else:
                    log.info(f"Response: {response}")
                    write_to_file(response)
                    XML.extract_and_print_values(file_path)
                    print(f"UPS Outlet Switched Group 3 Status: {XML.Binary_values}")
                    for i in range(len(XML.Binary_values)):
                        if XML.Binary_values[i] == '1':
                            global SOG3_Status
                            SOG3_Status = Microlink_functions.UnSwitched_OutletGroup_Status[i]
                            print(f"UPS Switched Outlet Group 3 Status: {Microlink_functions.UnSwitched_OutletGroup_Status[i]}")

                    break
            except Exception as ex:
                log.error(f"Exception caught: {ex}")
                retries += 1
                log.info(f"Retrying connection... Attempt {retries}")
                self.connect_to_simulator()
            time.sleep(5)



        if retries == max_retries:
            log.error("Max retries reached. Command failed.")


    #
    # def shutdown_microlink_simulator(self):
    #     try:
    #         self.tn.write(b'disconnect\n')
    #         self.tn.close()
    #         log.info("Simulator disconnected and closed")
    #     except Exception as ex:
    #         log.error(f"Error during shutdown: {ex}")











