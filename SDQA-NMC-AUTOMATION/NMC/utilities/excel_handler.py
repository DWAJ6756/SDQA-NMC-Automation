import os
from openpyxl import Workbook
from datetime import datetime


class ExcelHandler:

    def __init__(self, workbook):
        self.wb = workbook
        self.ws = workbook.active
        self.current_row = 2
        # Create test_results directory if it doesn't exist
        self.results_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'test_results')
        os.makedirs(self.results_dir, exist_ok=True)

    def write_test_data(self, test_name, nmc_value, simulator_value,operation_type, operation="READ-WRITE"):
        """Write test data to Excel sheet"""
        self.ws[f'A{self.current_row}'] = test_name
        self.ws[f'B{self.current_row}'] = nmc_value
        self.ws[f'C{self.current_row}'] = simulator_value
        self.ws[f'D{self.current_row}'] = operation
        self.ws[f'E{self.current_row}'] = operation_type
        self.current_row += 1

    def save(self, filename=None):
        """Save the workbook in the test_results directory"""
        if filename is None:
            filename = f'test_results_{datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}.xlsx'

        # Create full path for the file
        file_path = os.path.join(self.results_dir, filename)
        self.wb.save(file_path)